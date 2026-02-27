import datetime as dt
import io
import os
import re
import shutil
import time
import atexit
import zipfile
from typing import Dict, List, Optional
from xml.sax.saxutils import escape as _xml_escape

import pandas as pd
import sqlalchemy as sa
from sqlalchemy import event
from openpyxl import load_workbook, Workbook


DEFAULT_SERVER = "ANVDEVSQLVPM01"
DEFAULT_DB = "Aurora_APAC_DEV_Output_China"
DEFAULT_TEMPLATE = "China BKT template.xlsx"
DEFAULT_TABLES = [
    "dbo.ResourceGroupYear1",
    "dbo.ResourceGroupMonth1",
    "dbo.ResourceYear1",
    "dbo.ResourceMonth1",
    "dbo.ZoneHour1",
    "dbo.ZoneYear1",
    "dbo.LinkYear1",
]

# Global variables for open workbook
_open_workbook = None
_open_workbook_path = None
_open_worksheet = None
ZONEHOUR_NAMES = [
    "Anhui",
    "Beijing",
    "Chongqing",
    "Fujian",
    "Gansu",
    "Guangdong",
    "Guangxi",
    "Guizhou",
    "Hainan",
    "Hebei",
    "Heilongjiang",
    "Henan",
    "Hong Kong",
    "Hubei",
    "Hunan",
    "Inner Mongolia",
    "Jiangsu",
    "Jiangxi",
    "Jilin",
    "Liaoning",
    "Ningxia",
    "Qinghai",
    "Shaanxi",
    "Shandong",
    "Shanghai",
    "Shanxi",
    "Sichuan",
    "Tianjin",
    "Xinjiang",
    "Xizang",
    "Yunnan",
    "Zhejiang",
]
TEMPLATE_TARGETS = {
    "dbo.ResourceGroupYear1": ("ResourceGroupYear", 5),
    "dbo.ResourceYear1": ("ResourceYear", 4),
    "dbo.ResourceGroupMonth1": ("ResourceGroupMonth", 5),
    "dbo.ResourceMonth1": ("ResourceMonth", 4),
    "dbo.ZoneHour1": ("ZoneHour", 1),
    "dbo.ZoneYear1": ("ZoneYear", 1),
    "dbo.LinkYear1": ("LinkYear", 1),
}

# Columns to SELECT and their exact Excel column positions in the original template.
# Formula cols (A-C/D) are left untouched. Data written to original positions only.
# All other columns are cleared (left empty). Cross-sheet SUMIFS still reference original cols.
TABLE_COLUMN_POSITIONS = {
    "dbo.ResourceGroupYear1": {
        "Name": 6,         # F  (formula cols A-C reference this)
        "Time_Period": 8,  # H
        "Capacity": 10,    # J
        "Output_MWH": 17,  # Q
    },
    "dbo.ResourceYear1": {
        "Time_Period": 7,      # G
        "Zone": 8,             # H
        "Generation_MWh": 66,  # BN
        "Primary_Fuel": 72,    # BT  (formula col A references this)
    },
    "dbo.ResourceGroupMonth1": {
        "Name": 6,          # F  (formula cols A-C reference this)
        "Output_MWH": 17,   # Q
        "Report_Year": 39,  # AM
        "Report_Month": 40, # AN
    },
    "dbo.ResourceMonth1": {
        "Zone": 8,            # H
        "Generation_MWh": 66, # BN
        "Primary_Fuel": 72,   # BT  (formula col A references this)
        "Report_Year": 74,    # BV
        "Report_Month": 75,   # BW
    },
}

# Mapping from SQL table logical name → zip path inside xlsx
RESOURCE_SHEET_ZIP_MAP = {
    "dbo.ResourceGroupYear1":  "xl/worksheets/sheet10.xml",
    "dbo.ResourceYear1":       "xl/worksheets/sheet11.xml",
    "dbo.ResourceGroupMonth1": "xl/worksheets/sheet12.xml",
    "dbo.ResourceMonth1":      "xl/worksheets/sheet13.xml",
}

# Per-sheet configuration for the fast zip-based writer.
# formula_cells: (col_letter, formula_with_{r}_placeholder, t_attr)
# data_cols:     (sql_col_name, col_letter, is_string_type)
RESOURCE_SHEET_WRITE_CONFIG = {
    "dbo.ResourceGroupYear1": {
        "formula_cells": [
            ("A", 'LEFT(F{r},IFERROR(FIND("_",F{r},FIND("_",F{r},1)+1),FIND("_",F{r},1))-1)', "str"),
            ("B", "VLOOKUP(A{r},[7]Mapping!$A$2:$B$18,2,0)", "str"),
            ("C", 'RIGHT(F{r},LEN(F{r})-IFERROR(FIND("_",F{r},FIND("_",F{r},1)+1),FIND("_",F{r},1)))', "str"),
        ],
        "data_cols": [
            ("Name",        "F", True),
            ("Time_Period", "H", False),
            ("Capacity",    "J", False),
            ("Output_MWH",  "Q", False),
        ],
    },
    "dbo.ResourceYear1": {
        "formula_cells": [
            ("A", 'IFERROR(IF(BT{r}="Sun_Dist",BT{r},RIGHT(BT{r},LEN(BT{r})-FIND("_",BT{r},1))),BT{r})', "str"),
            ("B", "VLOOKUP(A{r},[7]Mapping!$A:$B,2,0)", "str"),
        ],
        "data_cols": [
            ("Time_Period",    "G",  False),
            ("Zone",           "H",  True),
            ("Generation_MWh", "BN", False),
            ("Primary_Fuel",   "BT", True),
        ],
    },
    "dbo.ResourceGroupMonth1": {
        "formula_cells": [
            ("A", 'LEFT(F{r},IFERROR(FIND("_",F{r},FIND("_",F{r},1)+1),FIND("_",F{r},1))-1)', "str"),
            ("B", "VLOOKUP(A{r},[7]Mapping!$A$2:$B$18,2,0)", "str"),
            ("C", 'RIGHT(F{r},LEN(F{r})-IFERROR(FIND("_",F{r},FIND("_",F{r},1)+1),FIND("_",F{r},1)))', "str"),
        ],
        "data_cols": [
            ("Name",         "F",  True),
            ("Output_MWH",   "Q",  False),
            ("Report_Year",  "AM", False),
            ("Report_Month", "AN", False),
        ],
    },
    "dbo.ResourceMonth1": {
        "formula_cells": [
            ("A", 'IFERROR(IF(BT{r}="Sun_Dist",BT{r},RIGHT(BT{r},LEN(BT{r})-FIND("_",BT{r},1))),BT{r})', "str"),
            ("B", "VLOOKUP(A{r},[7]Mapping!$A:$B,2,0)", "str"),
        ],
        "data_cols": [
            ("Zone",           "H",  True),
            ("Generation_MWh", "BN", False),
            ("Primary_Fuel",   "BT", True),
            ("Report_Year",    "BV", False),
            ("Report_Month",   "BW", False),
        ],
    },
}


def _col_to_int(col: str) -> int:
    """Convert Excel column letter to sort integer (A=1, Z=26, AA=27, ...)."""
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _gen_row_bytes(
    row_num: int,
    formula_cells: list,
    data_cols: list,
    row_dict: dict,
) -> bytes:
    """Generate XML bytes for one xlsx data row."""
    cells: dict[int, str] = {}

    for col_letter, formula_tmpl, cell_type in formula_cells:
        formula = formula_tmpl.format(r=row_num)
        col_int = _col_to_int(col_letter)
        cells[col_int] = f'<c r="{col_letter}{row_num}" t="{cell_type}"><f>{formula}</f></c>'

    for sql_col, col_letter, is_str in data_cols:
        val = row_dict.get(sql_col)
        if val is None:
            continue
        if isinstance(val, float) and val != val:  # NaN
            continue
        col_int = _col_to_int(col_letter)
        if is_str:
            escaped = _xml_escape(str(val))
            cells[col_int] = f'<c r="{col_letter}{row_num}" t="inlineStr"><is><t>{escaped}</t></is></c>'
        else:
            cells[col_int] = f'<c r="{col_letter}{row_num}"><v>{val}</v></c>'

    body = "".join(v for _, v in sorted(cells.items()))
    return f'<row r="{row_num}">{body}</row>\n'.encode("utf-8")


def _write_resource_sheets_fast(
    engine: sa.Engine,
    source_path: str,
    out_path: str,
    tables: List[str],
    year_start: int | None,
    year_end: int | None,
    chunksize: int = 50000,
) -> None:
    """Replace resource-sheet XMLs directly inside the xlsx zip.

    Bypasses openpyxl entirely for the 4 large resource sheets, saving ~90s of
    XML parsing.  Formula cells A-C are written as formula strings so Excel
    recalculates them on open; data columns are written as plain values.

    source_path: template (first export) or the existing output file (re-export).
    out_path:    destination xlsx (may equal source_path for in-place update).
    """
    # Step 1: fetch SQL data for all requested resource tables
    table_dfs: dict[str, "pd.DataFrame"] = {}
    for logical in tables:
        if logical not in RESOURCE_SHEET_WRITE_CONFIG:
            continue
        cfg = RESOURCE_SHEET_WRITE_CONFIG[logical]
        schema, tname = logical.split(".")
        full_name = f"[{schema}].[{tname}]"
        cols_sql = ", ".join(f'[{c}]' for c, _, _ in cfg["data_cols"])
        where_parts: list[str] = []
        params = None
        if year_start is not None and year_end is not None:
            where_parts.append("[Report_Year] BETWEEN ? AND ?")
            params = (year_start, year_end)
        where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        query = f"SELECT {cols_sql} FROM {full_name}{where_sql}"

        print(f"  SQL {logical} ...", end=" ", flush=True)
        t0 = time.perf_counter()
        with engine.connect().execution_options(stream_results=True) as conn:
            df = pd.read_sql_query(query, conn, params=params)
        print(f"{len(df)} rows in {time.perf_counter() - t0:.1f}s")
        table_dfs[logical] = df

    # Step 2: rebuild xlsx zip, replacing resource-sheet XMLs
    resource_zip_files = {
        RESOURCE_SHEET_ZIP_MAP[lg]: lg
        for lg in tables
        if lg in RESOURCE_SHEET_ZIP_MAP and lg in table_dfs
    }

    print("  Rebuilding xlsx ...", end=" ", flush=True)
    t0 = time.perf_counter()
    tmp_path = out_path + "._tmp"
    try:
        with zipfile.ZipFile(source_path, "r") as zin, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                fname = item.filename

                # Drop calcChain — Excel rebuilds it on open
                if fname == "xl/calcChain.xml":
                    continue

                if fname not in resource_zip_files:
                    # Copy all other files unchanged
                    zout.writestr(item, zin.read(fname))
                    continue

                # Replace this resource sheet
                logical = resource_zip_files[fname]
                cfg = RESOURCE_SHEET_WRITE_CONFIG[logical]
                df = table_dfs[logical]

                original = zin.read(fname)

                # Extract the header row (row 1) verbatim from the template
                m = re.search(rb'<row r="1"[^>]*>.*?</row>', original, re.DOTALL)
                header_row = m.group(0) if m else b""

                # Find the boundaries of <sheetData>...</sheetData>
                sd_start = original.find(b"<sheetData")
                sd_end = original.rfind(b"</sheetData>") + len(b"</sheetData>")
                before = original[:sd_start]
                after = original[sd_end:]

                # Generate new <sheetData>
                buf = io.BytesIO()
                buf.write(b"<sheetData>")
                if header_row:
                    buf.write(header_row)
                    buf.write(b"\n")
                formula_cells = cfg["formula_cells"]
                data_cols = cfg["data_cols"]
                for i, row_t in enumerate(df.itertuples(index=False), start=2):
                    buf.write(_gen_row_bytes(i, formula_cells, data_cols, row_t._asdict()))
                buf.write(b"</sheetData>")

                new_content = before + buf.getvalue() + after
                zout.writestr(fname, new_content)
                print(f"\n    {logical}: {len(df)} rows", end="", flush=True)

        if os.path.exists(out_path) and out_path != source_path:
            os.remove(out_path)
        if os.path.exists(out_path):
            os.remove(out_path)
        os.rename(tmp_path, out_path)
        print(f"\n  Zip done in {time.perf_counter() - t0:.1f}s")
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# xlsx zip path for the ZoneHour sheet (receives the full province pivot)
ZONEHOUR_ZIP_FILE = "xl/worksheets/sheet14.xml"


def _int_to_col(n: int) -> str:
    """Convert 1-based column index to Excel column letter (1→A, 28→AB, ...)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _write_zonehour_pivot_fast(
    engine: sa.Engine,
    source_path: str,
    out_path: str,
    year_start: int | None,
    year_end: int | None,
) -> None:
    """Write ZoneHour pivot to the ZoneHour sheet via direct zip XML replacement.

    Output columns: A=Report_Month, B=Report_Hour, C=Anhui, D=Beijing, ..., AH=Zhejiang
    (province order follows ZONEHOUR_NAMES).
    """
    price_cols_sql = [
        f"MAX(CASE WHEN [Name] = '{n}' THEN [Price] END) AS [{n}]"
        for n in ZONEHOUR_NAMES
    ]
    where_parts = ["[Condition] = 'Average'"]
    params = None
    if year_start is not None and year_end is not None:
        where_parts.append("[Report_Year] BETWEEN ? AND ?")
        params = (year_start, year_end)
    where_sql = " WHERE " + " AND ".join(where_parts)
    query = (
        "SELECT [Time_Period], [Report_Month], [Report_Hour], "
        + ", ".join(price_cols_sql)
        + " FROM [dbo].[ZoneHour1]"
        + where_sql
        + " GROUP BY [Time_Period], [Report_Month], [Report_Hour]"
        + " ORDER BY [Time_Period], [Report_Month], [Report_Hour]"
    )

    print("  SQL dbo.ZoneHour1 (pivot) ...", end=" ", flush=True)
    t0 = time.perf_counter()
    with engine.connect().execution_options(stream_results=True) as conn:
        df = pd.read_sql_query(query, conn, params=params)
    print(f"{len(df)} rows in {time.perf_counter() - t0:.1f}s")

    # Column layout: enumerate DataFrame columns → (name, col_letter, is_str)
    col_info = [
        (col_name, _int_to_col(i + 1), df[col_name].dtype == object)
        for i, col_name in enumerate(df.columns)
    ]

    def _header_row_bytes() -> bytes:
        cells = "".join(
            f'<c r="{col_letter}1" t="inlineStr"><is><t>{_xml_escape(col_name)}</t></is></c>'
            for col_name, col_letter, _ in col_info
        )
        return f'<row r="1">{cells}</row>\n'.encode("utf-8")

    print("  Rebuilding xlsx (ZoneHour pivot) ...", end=" ", flush=True)
    t0 = time.perf_counter()
    tmp_path = out_path + "._tmp"
    try:
        with zipfile.ZipFile(source_path, "r") as zin, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                fname = item.filename
                if fname == "xl/calcChain.xml":
                    continue
                if fname != ZONEHOUR_ZIP_FILE:
                    zout.writestr(item, zin.read(fname))
                    continue

                original = zin.read(fname)
                sd_start = original.find(b"<sheetData")
                sd_end = original.rfind(b"</sheetData>") + len(b"</sheetData>")
                before = original[:sd_start]
                after = original[sd_end:]

                buf = io.BytesIO()
                buf.write(b"<sheetData>")
                buf.write(_header_row_bytes())
                for i, row_t in enumerate(df.itertuples(index=False), start=2):
                    buf.write(_gen_row_bytes(i, [], col_info, row_t._asdict()))
                buf.write(b"</sheetData>")

                zout.writestr(fname, before + buf.getvalue() + after)
                print(f"\n    ZoneHour: {len(df)} rows", end="", flush=True)

        if os.path.exists(out_path):
            os.remove(out_path)
        os.rename(tmp_path, out_path)
        print(f"\n  Zip done in {time.perf_counter() - t0:.1f}s")
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


def _write_other_tables_fast(
    engine,
    source_path: str,
    out_path: str,
    table_infos: list,  # list of (logical, sheet_name, query, params)
) -> None:
    """Append other tables as new sheets directly into the xlsx zip (no openpyxl)."""
    table_dfs = []
    for logical, sheet_name, query, params in table_infos:
        print(f"  SQL {logical} ...", end=" ", flush=True)
        t0 = time.perf_counter()
        with engine.connect().execution_options(stream_results=True) as conn:
            df = pd.read_sql_query(query, conn, params=params)
        print(f"{len(df)} rows in {time.perf_counter() - t0:.1f}s")
        if not df.empty:
            table_dfs.append((sheet_name, df))

    if not table_dfs:
        return

    print("  Rebuilding xlsx (other tables) ...", end=" ", flush=True)
    t0 = time.perf_counter()
    tmp_path = out_path + "._tmp"
    try:
        with zipfile.ZipFile(source_path, "r") as zin, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            wb_xml  = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            ct_xml  = zin.read("[Content_Types].xml").decode("utf-8")

            next_sid  = max((int(m) for m in re.findall(r'sheetId="(\d+)"',        wb_xml)),  default=0) + 1
            next_rid  = max((int(m) for m in re.findall(r'Id="rId(\d+)"',          rels_xml)), default=0) + 1
            next_fnum = max((int(m) for m in re.findall(r'worksheets/sheet(\d+)\.xml', rels_xml)), default=0) + 1

            for item in zin.infolist():
                if item.filename not in ("xl/workbook.xml", "xl/_rels/workbook.xml.rels", "[Content_Types].xml"):
                    zout.writestr(item, zin.read(item.filename))

            new_wb_parts  = []
            new_rel_parts = []
            new_ct_parts  = []

            for sheet_name, df in table_dfs:
                col_info = [
                    (col, _int_to_col(i + 1), df[col].dtype == object)
                    for i, col in enumerate(df.columns)
                ]
                rid       = f"rId{next_rid}"
                file_path = f"worksheets/sheet{next_fnum}.xml"

                buf = io.BytesIO()
                buf.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
                buf.write(b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')
                header_cells = "".join(
                    f'<c r="{cl}1" t="inlineStr"><is><t>{_xml_escape(cn)}</t></is></c>'
                    for cn, cl, _ in col_info
                )
                buf.write(f'<row r="1">{header_cells}</row>\n'.encode("utf-8"))
                for i, row_t in enumerate(df.itertuples(index=False), start=2):
                    buf.write(_gen_row_bytes(i, [], col_info, row_t._asdict()))
                buf.write(b'</sheetData></worksheet>')

                zout.writestr(f"xl/{file_path}", buf.getvalue())
                print(f"\n    {sheet_name}: {len(df)} rows", end="", flush=True)

                new_wb_parts.append(
                    f'<sheet name="{_xml_escape(sheet_name)}" sheetId="{next_sid}" r:id="{rid}"/>'
                )
                new_rel_parts.append(
                    f'<Relationship Id="{rid}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                    f'Target="{file_path}"/>'
                )
                new_ct_parts.append(
                    f'<Override PartName="/xl/{file_path}" '
                    f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                )
                next_sid  += 1
                next_rid  += 1
                next_fnum += 1

            wb_xml   = wb_xml.replace("</sheets>",        "".join(new_wb_parts)  + "</sheets>")
            rels_xml = rels_xml.replace("</Relationships>", "".join(new_rel_parts) + "</Relationships>")
            ct_xml   = ct_xml.replace("</Types>",          "".join(new_ct_parts)  + "</Types>")

            zout.writestr("xl/workbook.xml",          wb_xml.encode("utf-8"))
            zout.writestr("xl/_rels/workbook.xml.rels", rels_xml.encode("utf-8"))
            zout.writestr("[Content_Types].xml",       ct_xml.encode("utf-8"))

        os.replace(tmp_path, out_path)
        print(f"\n  Zip done in {time.perf_counter() - t0:.1f}s")
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


def _enable_fast_executemany(engine: sa.Engine) -> None:
    @event.listens_for(engine, "before_cursor_execute")
    def _set_fast_executemany(conn, cursor, statement, parameters, context, executemany):
        if executemany:
            try:
                cursor.fast_executemany = True
            except Exception:
                pass


def _make_engine(server: str, database: str) -> sa.Engine:
    params = sa.engine.URL.create(
        "mssql+pyodbc",
        query={
            "odbc_connect": (
                "DRIVER={ODBC Driver 17 for SQL Server};"
                f"SERVER={server};"
                f"DATABASE={database};"
                "Trusted_Connection=Yes"
            )
        },
    )
    engine = sa.create_engine(params, fast_executemany=True)
    _enable_fast_executemany(engine)
    return engine


def _list_tables(engine: sa.Engine, database: str) -> List[Dict[str, str]]:
    sql = """
    SELECT TABLE_SCHEMA, TABLE_NAME
    FROM INFORMATION_SCHEMA.TABLES
    WHERE TABLE_TYPE = 'BASE TABLE' AND TABLE_CATALOG = ?
    ORDER BY TABLE_SCHEMA, TABLE_NAME
    """
    return pd.read_sql_query(sql, engine, params=[(database,)]).to_dict("records")


def _list_columns(engine: sa.Engine, database: str, schema: str, table: str) -> List[str]:
    sql = """
    SELECT COLUMN_NAME
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_CATALOG = ? AND TABLE_SCHEMA = ? AND TABLE_NAME = ?
    ORDER BY ORDINAL_POSITION
    """
    rows = pd.read_sql_query(sql, engine, params=[(database, schema, table)]).to_dict("records")
    return [r["COLUMN_NAME"] for r in rows]


def _safe_sheet_name(name: str, used: Dict[str, int]) -> str:
    # Excel sheet names: max 31 chars, no []:*?/\
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    base = cleaned[:31] if len(cleaned) > 31 else cleaned
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    suffix = f"_{used[base]}"
    trimmed = base[: (31 - len(suffix))]
    return f"{trimmed}{suffix}"


def _write_table_to_csv(
    engine: sa.Engine,
    database: str,
    schema: str,
    name: str,
    query: str,
    params,
    chunksize: int,
    tmp_csv: str,
    max_retries: int = 3,
) -> None:
    attempt = 0
    while True:
        attempt += 1
        if os.path.exists(tmp_csv):
            os.remove(tmp_csv)
        try:
            start = True
            with engine.connect().execution_options(stream_results=True) as conn:
                for chunk in pd.read_sql_query(query, conn, params=params, chunksize=chunksize):
                    chunk.to_csv(tmp_csv, index=False, header=start, mode="a", encoding="utf-8-sig")
                    start = False
            return
        except Exception as exc:
            if attempt > max_retries:
                raise
            print(
                f"Read failed for {schema}.{name} (attempt {attempt}/{max_retries}): {exc}. "
                f"Retrying in 5s with smaller chunksize."
            )
            time.sleep(5)
            chunksize = max(1000, chunksize // 2)


def _write_csv_to_sheet(writer: pd.ExcelWriter, sheet_name: str, tmp_csv: str, chunksize: int) -> None:
    """优化版：直接读取 CSV 并写入 Excel，减少 I/O 开销"""
    first = True
    startrow = 0
    with open(tmp_csv, 'r', encoding='utf-8-sig', newline='') as f:
        for line in f:
            if first:
                # 写入表头
                row_data = line.strip().split(',')
                for col_idx, value in enumerate(row_data):
                    col = col_idx + 1
                    if col < 1:
                        col = 1
                    writer.sheets[sheet_name].cell(row=startrow, column=col, value=value)
                first = False
                startrow += 1
            else:
                # 写入数据行
                row_data = line.strip().split(',')
                for col_idx, value in enumerate(row_data):
                    col = col_idx + 1
                    if col < 1:
                        col = 1
                    if startrow >= 0:
                        writer.sheets[sheet_name].cell(row=startrow, column=col, value=value)
                startrow += 1


def _write_table_to_csv_optimized(
    engine: sa.Engine,
    database: str,
    schema: str,
    name: str,
    query: str,
    params,
    chunksize: int,
    tmp_csv: str,
    max_retries: int = 3,
) -> None:
    """优化版：减少文件打开/关闭次数"""
    attempt = 0
    while True:
        attempt += 1
        if os.path.exists(tmp_csv):
            os.remove(tmp_csv)
        
        try:
            with engine.connect().execution_options(stream_results=True) as conn:
                with open(tmp_csv, 'w', encoding='utf-8-sig', newline='') as f:
                    header_written = False
                    for chunk in pd.read_sql_query(query, conn, params=params, chunksize=chunksize):
                        chunk.to_csv(f, index=False, header=not header_written, mode='a')
                        header_written = True
            return
        except Exception as exc:
            if attempt > max_retries:
                raise
            print(
                f"Read failed for {schema}.{name} (attempt {attempt}/{max_retries}): {exc}. "
                f"Retrying in 5s with smaller chunksize."
            )
            time.sleep(5)
            chunksize = max(1000, chunksize // 2)


def _get_or_create_workbook(out_path: str) -> tuple[Workbook, Optional[str]]:
    """获取或创建打开的 Excel 工作簿，如果已打开则复用"""
    global _open_workbook, _open_workbook_path, _open_worksheet
    
    if _open_workbook is not None and _open_workbook_path == out_path:
        # 检查工作簿是否仍然有效
        try:
            _open_workbook.active
            return _open_workbook, _open_worksheet
        except Exception:
            print("Warning: Workbook file was closed externally, reopening...")
            _open_workbook = None
            _open_worksheet = None
    
    # 需要打开新工作簿
    if os.path.exists(out_path):
        # 复用现有文件
        _open_workbook = load_workbook(out_path)
    else:
        # 创建新工作簿
        _open_workbook = Workbook()
        # 删除默认创建的 sheet
        if 'Sheet' in _open_workbook.sheetnames:
            _open_workbook.remove(_open_workbook['Sheet'])
    
    _open_workbook_path = out_path
    _open_worksheet = None
    return _open_workbook, _open_worksheet


def _get_or_create_worksheet(wb: Workbook, sheet_name: str, start_col: int = 1) -> Workbook:
    """获取或创建工作表"""
    global _open_worksheet
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # 确保工作表至少有足够的行和列
    if ws.max_row < 1:
        ws.max_row = 1
    if ws.max_column < start_col:
        ws.max_column = start_col
    
    _open_worksheet = ws
    return ws


def _cleanup_workbook():
    """清理打开的工作簿"""
    global _open_workbook, _open_workbook_path, _open_worksheet
    
    if _open_workbook is not None:
        try:
            _open_workbook.close()
        except Exception as e:
            print(f"Warning: Error closing workbook: {e}")
        _open_workbook = None
        _open_workbook_path = None
        _open_worksheet = None


def _ensure_template_output(out_path: str, template_path: str) -> None:
    """确保输出文件存在，如果存在模板则复制"""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    if os.path.exists(template_path):
        shutil.copyfile(template_path, out_path)
    else:
        wb = Workbook()
        wb.save(out_path)


def _parse_year_filter(year_arg: str):
    if not year_arg:
        return None, None
    if "-" in year_arg:
        parts = [p.strip() for p in year_arg.split("-", 1)]
        if len(parts) != 2 or not parts[0] or not parts[1]:
            raise ValueError("Invalid year range. Use YYYY or YYYY-YYYY.")
        return int(parts[0]), int(parts[1])
    return int(year_arg), int(year_arg)


def _ensure_template_output(out_path: str, template_path: str) -> None:
    d = os.path.dirname(out_path)
    if d:
        os.makedirs(d, exist_ok=True)
    if os.path.exists(template_path):
        shutil.copyfile(template_path, out_path)
        return
    wb = Workbook()
    wb.save(out_path)


def _clear_sheet_from_col(ws, start_col: int) -> None:
    """只清除指定列开始的数据，不删除整个工作表"""
    # 清除 start_col 列及之后的所有列
    if ws.max_row == 0 or ws.max_column == 0:
        return
    for row_idx in range(1, ws.max_row + 1):
        for col in range(start_col, ws.max_column + 1):
            try:
                cell = ws.cell(row=row_idx, column=col)
                cell.value = None
            except ValueError:
                continue


def _write_positioned_to_sheet(
    engine: sa.Engine,
    query: str,
    params,
    chunksize: int,
    out_path: str,
    sheet_name: str,
    col_positions: dict,
    clear_from_col: int,
    max_retries: int = 3,
    save: bool = True,
) -> None:
    """Write SQL result columns to specific (non-sequential) Excel column positions.
    Formula cols (before clear_from_col) are untouched. All data rows (row 2+) are
    cleared with delete_rows() for speed, then new data is written cell-by-cell.
    col_positions: {sql_col_name: excel_col_index (1-based)}
    Set save=False to skip saving (caller will save once at the end).
    """
    wb, _ = _get_or_create_workbook(out_path)
    ws = _get_or_create_worksheet(wb, sheet_name, clear_from_col)

    # Fast clear: delete all data rows at once (much faster than cell-by-cell)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Write column headers to row 1 (overwrite only target positions)
    for col_name, col_idx in col_positions.items():
        ws.cell(row=1, column=col_idx, value=col_name)

    attempt = 0
    while True:
        attempt += 1
        current_row = 2
        try:
            with engine.connect().execution_options(stream_results=True) as conn:
                for chunk in pd.read_sql_query(query, conn, params=params, chunksize=chunksize):
                    for row_tuple in chunk.itertuples(index=False):
                        row_dict = row_tuple._asdict()
                        for col_name, col_idx in col_positions.items():
                            val = row_dict.get(col_name)
                            if val is not None:
                                ws.cell(row=current_row, column=col_idx, value=val)
                        current_row += 1
            break
        except Exception as exc:
            if attempt > max_retries:
                raise
            print(f"Read failed (attempt {attempt}/{max_retries}): {exc}. Retrying in 5s.")
            time.sleep(5)

    if save:
        wb.save(out_path)
    print(f"  -> {current_row - 2} rows written to '{sheet_name}'")


def _write_csv_to_sheet_offset(out_path: str, sheet_name: str, tmp_csv: str, chunksize: int, start_col: int) -> None:
    """优化版：使用已打开的工作簿进行写入"""
    if not start_col or start_col < 1:
        start_col = 1
    wb, ws = _get_or_create_workbook(out_path)
    ws = _get_or_create_worksheet(wb, sheet_name, start_col)
    
    # 只清除指定列开始的数据，不删除整个工作表
    if ws.max_row > 0 and ws.max_column >= start_col:
        for row_idx in range(1, ws.max_row + 1):
            for col in range(start_col, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = None
                except ValueError:
                    continue
    
    wb.save(out_path)
    
    # 使用优化后的 CSV 读取函数
    with open(tmp_csv, 'r', encoding='utf-8-sig', newline='') as f:
        first = True
        startrow = 1
        for line in f:
            if first:
                # 写入表头
                row_data = line.strip().split(',')
                for col_idx, value in enumerate(row_data):
                    col = start_col + col_idx
                    if col < 1:
                        col = 1
                    if startrow >= 1 and col >= 1:
                        ws.cell(row=startrow, column=col, value=value)
                first = False
                startrow += 1
            else:
                # 写入数据行
                row_data = line.strip().split(',')
                for col_idx, value in enumerate(row_data):
                    col = start_col + col_idx
                    if col < 1:
                        col = 1
                    if startrow >= 1 and col >= 1:
                        ws.cell(row=startrow, column=col, value=value)
                startrow += 1
    wb.save(out_path)


def export_database(
    server: str,
    database: str,
    out_path: str,
    chunksize: int,
    tables: List[str],
    year_start: int | None,
    year_end: int | None,
) -> None:
    engine = _make_engine(server, database)
    all_tables = _list_tables(engine, database)
    if not all_tables:
        print(f"No tables found in {database}.")
        return
    wanted = set(tables)
    tables = [t for t in all_tables if f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" in wanted]
    if not tables:
        print("None of the requested tables were found.")
        return

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    used_sheet_names: Dict[str, int] = {}

    tmp_dir = os.path.join(os.path.dirname(out_path), "_tmp_exports")
    os.makedirs(tmp_dir, exist_ok=True)

    # 使用已打开的工作簿
    wb, _ = _get_or_create_workbook(out_path)
    
    for tbl in tables:
        schema = tbl["TABLE_SCHEMA"]
        name = tbl["TABLE_NAME"]
        full_name = f"[{schema}].[{name}]"
        sheet_name = _safe_sheet_name(f"{schema}.{name}", used_sheet_names)

        print(f"Exporting {full_name} -> sheet '{sheet_name}'")
        query = f"SELECT * FROM {full_name}"
        params = None
        year_filter_sql = ""
        year_params = None
        if year_start is not None and year_end is not None:
            year_filter_sql = " WHERE [Report_Year] BETWEEN ? AND ?"
            year_params = (year_start, year_end)

        if f"{schema}.{name}" == "dbo.ZoneHour1":
            select_cols = [
                "Time_Period",
                "Report_Month",
                "Report_Hour",
            ]
            price_cols = [
                f"MAX(CASE WHEN [Name] = '{n}' THEN [Price] END) AS [{n}]"
                for n in ZONEHOUR_NAMES
            ]
            where_parts = ["[Condition] = 'Average'"]
            if year_filter_sql:
                where_parts.append("[Report_Year] BETWEEN ? AND ?")
                params = year_params
            where_sql = " WHERE " + " AND ".join(where_parts)
            query = (
                "SELECT "
                + ", ".join(select_cols + price_cols)
                + f" FROM {full_name}"
                + where_sql
                + " GROUP BY [Time_Period], [Report_Month], [Report_Hour]"
                + " ORDER BY [Time_Period], [Report_Month], [Report_Hour]"
            )
        else:
            if year_filter_sql:
                query = f"SELECT * FROM {full_name}{year_filter_sql}"
                params = year_params

        tmp_csv = os.path.join(tmp_dir, f"{schema}.{name}.csv")
        _write_table_to_csv_optimized(
            engine=engine,
            database=database,
            schema=schema,
            name=name,
            query=query,
            params=params,
            chunksize=chunksize,
            tmp_csv=tmp_csv,
        )
        _write_csv_to_sheet_offset(out_path, sheet_name, tmp_csv, chunksize, 1)

    wb.save(out_path)
    print(f"Done. Saved: {out_path}")


def export_to_template(
    server: str,
    database: str,
    out_path: str,
    chunksize: int,
    tables: List[str],
    year_start: int | None,
    year_end: int | None,
) -> None:
    engine = _make_engine(server, database)
    all_tables = _list_tables(engine, database)
    if not all_tables:
        print(f"No tables found in {database}.")
        return
    wanted = set(tables)
    found_tables = [t for t in all_tables if f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" in wanted]
    if not found_tables:
        print("None of the requested tables were found.")
        return

    # Classify tables by write strategy
    resource_logicals = [
        f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}"
        for t in found_tables
        if f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" in RESOURCE_SHEET_WRITE_CONFIG
    ]
    zonehour_present = any(
        f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" == "dbo.ZoneHour1"
        for t in found_tables
    )
    other_tables = [
        t for t in found_tables
        if f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" not in RESOURCE_SHEET_WRITE_CONFIG
        and f"{t['TABLE_SCHEMA']}.{t['TABLE_NAME']}" != "dbo.ZoneHour1"
    ]

    # Close any open openpyxl handle before zip operations replace the file on disk
    if resource_logicals or zonehour_present:
        global _open_workbook, _open_workbook_path, _open_worksheet
        if _open_workbook is not None:
            try:
                _open_workbook.close()
            except Exception:
                pass
            _open_workbook = None
            _open_workbook_path = None
            _open_worksheet = None

    # --- Fast path: 4 resource sheets via zip XML replacement ---
    if resource_logicals:
        _write_resource_sheets_fast(
            engine=engine,
            source_path=out_path,
            out_path=out_path,
            tables=resource_logicals,
            year_start=year_start,
            year_end=year_end,
            chunksize=chunksize,
        )

    # --- Fast path: ZoneHour pivot via zip XML replacement ---
    if zonehour_present:
        _write_zonehour_pivot_fast(
            engine=engine,
            source_path=out_path,
            out_path=out_path,
            year_start=year_start,
            year_end=year_end,
        )

    # --- Fast path: remaining tables appended as new sheets via zip ---
    if other_tables:
        year_params = (year_start, year_end) if year_start is not None else None
        where_sql = " WHERE [Report_Year] BETWEEN ? AND ?" if year_params else ""
        table_infos = []
        for tbl in other_tables:
            schema = tbl["TABLE_SCHEMA"]
            name   = tbl["TABLE_NAME"]
            logical = f"{schema}.{name}"
            sheet_name = TEMPLATE_TARGETS.get(logical, (name, 1))[0]
            query  = f"SELECT * FROM [{schema}].[{name}]{where_sql}"
            table_infos.append((logical, sheet_name, query, year_params))
        _write_other_tables_fast(engine, out_path, out_path, table_infos)

    print(f"Done. Updated: {out_path}")


def _prompt_years() -> tuple[int | None, int | None]:
    while True:
        raw = input("Set years (YYYY or YYYY-YYYY, blank for all): ").strip()
        if not raw:
            return None, None
        try:
            return _parse_year_filter(raw)
        except ValueError as exc:
            print(f"Invalid input: {exc}")

def _input_with_prefill(prompt: str, prefill: str) -> str:
    """Show input prompt with pre-filled text the user can edit in-place."""
    try:
        import readline
        readline.set_startup_hook(lambda: readline.insert_text(prefill))
        try:
            return input(prompt).strip()
        finally:
            readline.set_startup_hook()
    except (ImportError, AttributeError):
        # Windows fallback: simulate keystrokes via msvcrt
        try:
            import msvcrt
            import sys
            sys.stdout.write(prompt + prefill)
            sys.stdout.flush()
            buf = list(prefill)
            while True:
                ch = msvcrt.getwch()
                if ch in ("\r", "\n"):
                    sys.stdout.write("\n")
                    sys.stdout.flush()
                    return "".join(buf).strip()
                elif ch in ("\x08", "\x7f"):  # Backspace
                    if buf:
                        buf.pop()
                        sys.stdout.write("\b \b")
                        sys.stdout.flush()
                elif ch == "\x03":  # Ctrl+C
                    raise KeyboardInterrupt
                elif ch >= " ":
                    buf.append(ch)
                    sys.stdout.write(ch)
                    sys.stdout.flush()
        except ImportError:
            return input(f"{prompt}[{prefill}] ").strip() or prefill


def _prompt_database(default_db: str) -> str:
    raw = _input_with_prefill("Database name: ", default_db)
    return raw or default_db


def _prompt_out_path(default_path: str) -> str:
    raw = input("Output Excel full path (blank for default): ").strip()
    if not raw:
        return default_path
    if not raw.lower().endswith(".xlsx"):
        raw += ".xlsx"
    return raw


def main() -> None:
    server = DEFAULT_SERVER
    database = DEFAULT_DB
    chunksize = 50000
    year_start = None
    year_end = None
    out_path_override = ""

    while True:
        yr_label = "all" if year_start is None else f"{year_start}-{year_end}"
        out_label = out_path_override or "default"
        print("")
        print("Aurora Output Export (China)")
        print(f"Server: {server} | DB: {database} | Years: {yr_label} | Out: {out_label}")
        print("Options:")
        print("  a. set years")
        print("  b. set database name")
        print("  c. set output file")
        print("  x. export all tables")
        print("  q. quit")
        choice = input("Select option: ").strip().lower()

        if choice == "q":
            return
        if choice == "a":
            year_start, year_end = _prompt_years()
            continue
        if choice == "b":
            database = _prompt_database(database)
            continue
        if choice == "c":
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_out = os.path.join("exports", f"{database}_{ts}.xlsx")
            out_path_override = _prompt_out_path(default_out)
            continue
        if choice == "x":
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = out_path_override or os.path.join("exports", f"{database}_{ts}.xlsx")
            _ensure_template_output(out_path, DEFAULT_TEMPLATE)
            export_to_template(server, database, out_path, chunksize, DEFAULT_TABLES, year_start, year_end)
            continue

        print("Unknown option. Try again.")


if __name__ == "__main__":
    atexit.register(_cleanup_workbook)
    main()
